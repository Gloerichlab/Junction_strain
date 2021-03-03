

# -*- coding: utf-8 -*-
"""
@author: Jooske Monster
first used in publication: Monster et al. 2021 (JCB), https://doi.org/10.1083/jcb.202001042
disclaimer: I am not an experienced programmer so it will definitely not be the smartest script ever
feel free to use and edit script, but please refer to us accordingly :) 


Added additional annotations for clarity 03/03/2021

----------------------------------------------------------------------------------------------------------
PICTURE OF JUNCTIONAL STRAIN

input: segmented (skeletonized) images of junctions (outlines) 
(used SeedWaterSegmenter:https://pypi.org/project/SeedWaterSegmenter/ )
output: 
     excel files:
          Nodes.xlsx (with per frame: NodeID, NodeArray (4 px square coords of node), NodeCells (Nodes neighboring which cells), NodeIndex)
          Outlines.xlsx (with per frame: area (=length), coordinates, and both coords of centroids)
          Outlines_matched_xframes.xlsx (with per frame for both the junction in that frame (..2), and in the matched previous frame (...1): centroids of junctions, coordinates of junctions, area (=length), and for these combined: the difference in length (dLength) absolute, the relative difference in length (dLength strain), and lookup tables to indicate these) 
          
     individual images and/or stacks of:
          outlines
          nodes
          outlines_cut
          outlinesIID (after matching to previous frame, to be able to check how well they match)
          outlines_strain ([junction2-junction1]/junction1, percentage)
          outlines_absolute (note that this is in px)
          outlines_strain_capped (you can cap the amount of strain to make playing with the luts a bit easier)
          
  
1. opens individual images and makes stacks from them 
2. detects nodes (=tricellular junctions) where three junctions come together
3. subtracts the nodes from the segments to get individual segments (junctions)
4. using skimage.measure.regionprops detects individual segments (junctions) and extracts properties
5. Matching segments of frames:
     a.detects the centroids of segments and searches nearby centroids in radius (adjustable) 
     of previous frame using scipy.spatial.KDtree
     b.measures the (euclidian) distance between the centroid and all its nearby centroids of previous frame
     c. picks the shortest distance
          ones that do not find any (mostly when new junctions are formed or incorrectly segmented) are excluded
     d. creating new list where the junctions are matched between frames
6. calculates absolute and relative distances in junction length (=area as it is a skeletonized image)
7. creates 16-bit heatmap pictures where the intensity values correspond the difference in junction length
8. I used ImageJ to change LUTS and contrast, but feel free to write it yourself

"""


import numpy as np
from tifffile import imread, imwrite, TiffWriter
import pandas as pd
import glob
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook, Workbook
from skimage import io
from skimage.measure import regionprops_table
from scipy import spatial, ndimage
#from matplotlib import pyplot as plt
import math

#----------------------------------------------------------------------------
#settings you can play with
radius=50           #the radius it uses to match centroids between frames (in pixels)
deltaframes=2       #the number of frames between it calculates the change
cap=100             #the percentage strain above (or below) will just be the maximum color (value 100)

#----------------------------------------------------------------------------
#functions



def surrounding_pixels(x): 
     pixels = []
     pixels.append((x[0],x[1]))
     pixels.append((x[0]+1,x[1]))
     pixels.append((x[0],x[1]+1))
     pixels.append((x[0]+1,x[1]+1))
     return (pixels)

#searches where three or more objects intercept
def find_nodes(img): 
     img_nodes = np.zeros(img.shape, dtype=np.uint16)                           #empty picture for nodes to fill (16-bit)
     NodeID = []                                                                
     NodeArray = []                                                             
     NodeCells =[]   
     NodeIndex = []                                                          
     nodecounter = 0                                                           
     for index, x in np.ndenumerate(img):                                       #goes through all pixels
          if (index[0] < img.shape[0]-1) and (index[1] < img.shape[1]-1):       #to get rid of borders where error will occur
               a = surrounding_pixels(index)                                    #takes surrounding 2x2 pixels, and retrieves their values (=cell ID)
               pix1=img[a[0]]
               pix2=img[a[1]]
               pix3=img[a[2]]
               pix4=img[a[3]]
               pix1234=[pix1, pix2, pix3, pix4]    
               if len(set(pix1234))>2:                                          #if there are more than 2 unique values, it means that three cells intersect
                    NodeIndex.append(nodecounter)
                    nodecounter +=1
#                    NodeID.append("Node%03d_TCJ%s"%(nodecounter, list(set(pix1234))))                       #makes list of NodeIDs ("Node001")
                    NodeID.append("Node%03d"%nodecounter)
                    NodeArray.append(a)                                         #makes list of locations of nodes(2x2 array xy location)
                    NodeCells.append(list(set(pix1234)))                        #makes list of cells in the node (removes doubles with set)     
                    img_nodes[a[0]]=1                                           #it then fills the four pixels with 1 in the new image
                    img_nodes[a[1]]=1 
                    img_nodes[a[2]]=1 
                    img_nodes[a[3]]=1
#     plt.imshow(img_nodes)
     return img_nodes, NodeID, NodeArray, NodeCells, NodeIndex      

def surrounding_pixels2(x):
     pixels = []
     pixels.append((x[0]-1,x[1]-1))
     pixels.append((x[0]-1,x[1]))
     pixels.append((x[0]-1,x[1]+1))
     pixels.append((x[0],x[1]-1))
     pixels.append((x[0],x[1]))
     pixels.append((x[0],x[1]+1))
     pixels.append((x[0]+1,x[1]-1))
     pixels.append((x[0]+1,x[1]))
     pixels.append((x[0]+1,x[1]+1))
     return (pixels)

def calculateDistance(src, dest):  
     x1=src[0]
     x2=dest[0]
     y1=src[1]
     y2=dest[1]
     dist = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)  
     return dist

def tup(lst):
     lst2=lst.split("\n")   
     newlist=[]
     for i in lst2:
          j= i[0:6]
          j=j.replace(" ", "")
          j=j.replace("[", "")
          j=j.replace("]", "")
          k= i[6:]
          k=k.replace(" ", "")
          k=k.replace("[", "")
          k=k.replace("]", "") 
          l=(int(j), (int(k)))
          newlist.append(l)              
     return newlist


#-------------------------------------------------------------------------------
# 1. opening folder, and checking if already before stacks have been made

#popup window to ask for directory where the folders of the segmentation are in
root = tk.Tk()
root.withdraw()
print ("the popup window can be behind your terminal")
parent_dir=filedialog.askdirectory(title="open directory with segmentation folders")

#checks if the stacks and nodes are already made, and if not makes them

image_c_dir = parent_dir+"/Segments/*.tif"
image_j_dir=  parent_dir+"/Outlines/*.tif"

#-----------------------------------------------------------------------------
# 2. checks if the stacks and nodes are already made, and if not makes them
if os.path.isdir(parent_dir+"/Nodes")==False:    
     print ("finding nodes and making files...")
     
     #finds nodes using the function and makes stacks out of them as well as the segments and outlines files
     image_c_list=[]
     image_n_list=[]
     image_j_list=[]
     os.mkdir(os.path.join(parent_dir, "Nodes") )
     #preparations for excelfile (creates file and opens writer)
     excelpath = parent_dir+"/Nodes.xlsx" 
     Workbook().save(filename=excelpath)
     book = load_workbook(excelpath)
     writer = pd.ExcelWriter(excelpath, engine = 'openpyxl')
     writer.book = book
          
     for file in sorted(glob.glob(image_c_dir)): #loops over all the files in path
          print (file)
          read=imread(file)
          
          #finds nodes and adds image to stack and as a separate file in folder
          nodeimage=find_nodes(read)[0]      #this is the function to find the nodes      
          imwrite(parent_dir+"/Nodes/Node_"+file[-7:-4]+".tif", nodeimage)               
          image_c_list.append(read)          #adds cell images to stack
          image_n_list.append(nodeimage)     #adds node images to stack
          
          #gets properties of nodes and saves them to excel
          nodes_1=find_nodes(read)[1] #NodeID
          nodes_2=find_nodes(read)[2] #NodeArray
          nodes_3=find_nodes(read)[3] #NodeCells
          nodes_4=find_nodes(read)[4] #NodeIndex        
          Nodes=list(zip(nodes_1, nodes_2, nodes_3, nodes_4))
          df_nodes=pd.DataFrame(data=Nodes, columns= ["NodeID", "NodeArray", "NodeCells", "NodeIndex"])
          df_nodes.to_excel(writer, sheet_name = "Nodes_t"+file[-7:-4])
     writer.save()
     writer.close()
  
             
     for file in sorted(glob.glob(image_j_dir)): #loops over all the files in path
          read=imread(file)
          image_j_list.append(read)          #adds junction images to stack
     
     #saves the stacks to tiff
     image_c_stack = np.stack(image_c_list, axis=0)
     image_j_stack = np.stack(image_j_list, axis=0)
     image_n_stack = np.stack(image_n_list, axis=0)
     with TiffWriter(parent_dir+"/Segments_stack.tif") as tif:
         tif.save(image_c_stack)
     with TiffWriter(parent_dir+"/Outlines_stack.tif") as tif:
         tif.save(image_j_stack)
     with TiffWriter(parent_dir+"/Nodes_stack.tif") as tif:
         tif.save(image_n_stack)
     
     # makes stack of all stacks
     image_stack=np.stack([image_c_stack, image_j_stack, image_n_stack], axis=1)
     with TiffWriter(parent_dir+"/Merged_stack.tif", imagej=True) as tif:
         tif.save(image_stack)
     print ("files written: Segments_stack.tif, Outlines_stack.tif, Nodes_stack.tif, Merged_stack.tif, Nodes.xlsx, Nodes folder")
 
image_n_dir = parent_dir+"/Nodes/*.tif"

#-----------------------------------------------------------------------------
# 3. dilates the nodes and subtracts this from the outlines

#dilates the nodes
if os.path.isdir(parent_dir+"/Nodes_dilated")==False:    
     print ("dilating nodes...")
     os.mkdir(os.path.join(parent_dir, "Nodes_dilated") )
     for file in sorted(glob.glob(image_n_dir)): #loops over all the files in path
         read=imread(file)
     #     print (file)
         nodeimage_dilated = np.zeros(read.shape, dtype=np.uint16)
         for index, x in np.ndenumerate(read):                                  #goes through all pixels
              if (index[0] < read.shape[0]-1) and (index[1] < read.shape[1]-1): #to get rid of borders where error will occur
                   if x>0:                                                                #if pixel is positive 
                        a = surrounding_pixels2(index)                                    #takes surrounding 3x3 pixels
                        for i in a:
                             nodeimage_dilated[i]=1                                       #it then fills the pixels with 1 in the new image
         imwrite(parent_dir+"/Nodes_dilated/Nodes_dilated_"+file[-7:-4]+".tif", nodeimage_dilated)

image_n2_dir = parent_dir+"/Nodes_dilated/*"

# substract nodes from junction image to get isolated junctions
if os.path.isdir(parent_dir+"/Outlines_cut")==False:    
     print ("substracting nodes from junctions...")
     os.mkdir(os.path.join(parent_dir, "Outlines_cut") )
     index = 0
     for file in sorted(glob.glob(image_n2_dir)): #loops over all the files in path
          read_node = io.imread(file)
          read_outline = io.imread(sorted(glob.glob(image_j_dir))[index])
          cut = np.zeros(read_node.shape, dtype=np.uint16)
          tcj1=(read_node==1)
          junct1=(read_outline==1)
          cut[junct1]=1
          cut[tcj1]=0
          imwrite(parent_dir+"/Outlines_cut/Outlines_cut_"+file[-7:-4]+".tif", cut)
          index +=1


image_j2_dir = parent_dir+"/Outlines_cut/*.tif"

#-----------------------------------------------------------------------------
# 4. label the junctions and extract information

#prepares to write in excel
excelpath = parent_dir+"/Outlines.xlsx" 
Workbook().save(filename=excelpath)
book = load_workbook(excelpath)
writer = pd.ExcelWriter(excelpath, engine = 'openpyxl')
writer.book = book


s = [[1,1,1],[1,1,1],[1,1,1]] #structure parameter    
print ("detecting Outlines...")
for file in sorted(glob.glob(image_j2_dir)): #loops over cut Outlines
     read=io.imread(file) 
     #labels the junctions
     mask = read == 1
     label_mask, num_labels = ndimage.label(mask, structure = s) 
#     imwrite(parent_dir+"/Outlines_cut/Outlines_id_"+file[-7:-4]+".tif", np.uint16(label_mask))
#     junctionlabels = color.label2rgb(label_mask, bg_label=0)
#     plt.imshow(junctionlabels)
     #gets info from junctions
     props = regionprops_table(label_mask, properties=['label', 'area', 'coords', "centroid"])
     df1 = pd.DataFrame(props)
     df1.to_excel(writer, sheet_name = "Outline_t"+file[-7:-4])
     
#     centroids2=data2["centroids"].tolist()
print ("saved Outlines.xlsx")
writer.save()
writer.close()

#-----------------------------------------------------------------------------
# 5. matching centroids between two frames




excelpath = parent_dir+"/Outlines_matched_"+str(deltaframes)+"frames.xlsx" 
Workbook().save(filename=excelpath)
book = load_workbook(excelpath)
writer = pd.ExcelWriter(excelpath, engine = 'openpyxl')
writer.book = book


print ("matching centroids...")
for file in sorted(glob.glob(image_j2_dir))[0:-deltaframes]:   #-deltaframes because the last frame does not have a +1 frame 
     index1=file[-7:-4]
     index2=str((int(file[-7:-4])+deltaframes)).zfill(3)    #next sheet         
#     print (index1, index2)
     
     #reads the excelfile for centroids
     df1 = pd.read_excel(parent_dir+"/Outlines.xlsx" , sheet_name="Outline_t"+index1)
     df2 = pd.read_excel(parent_dir+"/Outlines.xlsx" , sheet_name="Outline_t"+index2)
     df1['centroids'] = df1[['centroid-0', 'centroid-1']].apply(tuple, axis=1)
     df2['centroids'] = df2[['centroid-0', 'centroid-1']].apply(tuple, axis=1)
     centroids1=df1["centroids"].tolist()
     centroids2=df2["centroids"].tolist()



     #5a. finds all neighbors of centroids 1 in radius x in centroids 2
     points = np.asarray(centroids1)
     tree = spatial.KDTree(points)
     sorted_centroids1 = []
     for results in tree.query_ball_point(centroids2, radius):                           #check if I have to make it query_ball_tree if too slow
         nearby_points = points[results]
#         print (nearby_points)                                                       #use this to find the optimal radius where most points have one other coord.
         sorted_centroids1.append(nearby_points)

     # 5b + c. finds closest centroid of centroids_sorted for each coordinate of centroids2     
     cleaned_sorted_centroids1=[]                                               #new list with one value per centroids2 coordinate
     index2_uncalled=[]
     for i in range(len(centroids2)):
          if len(sorted_centroids1[i])==0:                                      #if there is no nearby centroid
               index2_uncalled.append(i)                                        #puts the index of the coordinates in separate list
          else:                                                                 #if there is 1 or more values
               distances = []
               for j in sorted_centroids1[i]:                                   #loops over the possibly multiple coordinates close to centroids2
                    distances.append(calculateDistance(j, centroids2[i]))       #calculates the distances for each coordinate on the list
     #          print (i, distances)
               closest_index = min(range(len(distances)), key=distances.__getitem__) #finds the shortest distance index
     #          print (closest_index)
               cleaned_sorted_centroids1.append(sorted_centroids1[i][closest_index]) #and appends that coordinate to the new list
     
     
     #5d 
     #makes a tuple from the centroids
     cleaned_sorted_tuple_centroids1 =[]
     for i in range(len(cleaned_sorted_centroids1)):
          tupl=(cleaned_sorted_centroids1[i][0], cleaned_sorted_centroids1[i][1])
          cleaned_sorted_tuple_centroids1.append(tupl)

     #finds indexes of the matched centroids in the old data1 dataframe to retrieve other info in next part
     corrected_indexes_centroids1 =[]
     for i in range(len(cleaned_sorted_tuple_centroids1)):
     #    print (cleaned_sorted_tuple_centroids1[i])
          new_index=int(np.where(df1["centroids"] == cleaned_sorted_tuple_centroids1[i])[0])
     #    print(new_index)
          corrected_indexes_centroids1.append(new_index)
     
     #gets coords and area in order of the matched
     corrected_coords1= []
     corrected_area1=[]
     for i in corrected_indexes_centroids1:
     #     print (i)
          if isinstance(i, str)==True:
              corrected_coords1.append("NaN")
              corrected_area1.append("NaN")
          else:
               coord=df1["coords"][i]
               area=df1["area"][i]
               corrected_coords1.append(coord)
               corrected_area1.append(area) 
     
     #makes new list for area and coords in 2 with uncalled junctions deleted
     area2=list(df2["area"])
     coords2=list(df2["coords"])
     for i in range(len(index2_uncalled) - 1, -1, -1):                     #loop reversed so indexes do not shift
          j=index2_uncalled[i]
          del area2[j]
          del coords2[j]


#------------------------------------------------------------------------------------
#6 calculates difference in length between two dataframes
          
     dLength_absolute=[]
     dLength_strain=[]
     for i in range(len(corrected_area1)):         
          absolute = int(area2[i] - corrected_area1[i])
          strain = float(absolute/corrected_area1[i])     
          dLength_absolute.append(absolute)
          dLength_strain.append(strain)     
          

     #makes the pixel values to be used in the images later
          
     Bitmap1=[]
     for i in dLength_absolute:
          Bitmap1.append(i+1000)   # 1000 to make 1000 the 0 point to allow for negative values)  
     Bitmap2=[]
     for i in dLength_strain:
          Bitmap2.append(i*100+1000) #times 100 to get percentage (and higher res for bitmap), and 1000 to make 1000 the 0 point to allow for negative values)          

     Bitmap3=[]
     for i in Bitmap2:
          if i>(1000+cap): 
               Bitmap3.append(1000+cap)
          elif i<(1000-cap):
               Bitmap3.append(1000-cap)
          else:
               Bitmap3.append(i)

     
     #make new pandas dataframe
     newlist=list(zip(centroids2, cleaned_sorted_tuple_centroids1, coords2, corrected_coords1, 
                      area2, corrected_area1, dLength_absolute, dLength_strain, 
                      Bitmap1, Bitmap2, Bitmap3))   
     df3 = pd.DataFrame(data=newlist, columns= ["centroids2", "centroids1", "coords2", "coords1", 
                                                "area2", "area1", "dLength(absolute)", "dLength(strain)", 
                                                "Lookup table 16Bit absolute", "Lookup table 16Bit strain", "Lookup table 16Bit strain capped"])
     df3.to_excel(writer, sheet_name = "Outline_t"+index2)
writer.save()
writer.close()

print ("saved Outlines_matched_"+str(deltaframes)+"frames.xlsx")


#------------------------------------------------------------------------------------     
# 7. making the color-coded images (junction IDs to check overlap and differences between frames)

if os.path.isdir(parent_dir+"/Outlines_Id_"+str(deltaframes))==False:    
     os.mkdir(os.path.join(parent_dir, "Outlines_Id_"+str(deltaframes)) )
if os.path.isdir(parent_dir+"/Outlines_strain_"+str(deltaframes))==False:    
     os.mkdir(os.path.join(parent_dir, "Outlines_strain_"+str(deltaframes)) )    
if os.path.isdir(parent_dir+"/Outlines_strain_capped"+str(cap)+"_"+str(deltaframes))==False:    
     os.mkdir(os.path.join(parent_dir, "Outlines_strain_capped"+str(cap)+"_"+str(deltaframes)) )      
if os.path.isdir(parent_dir+"/Outlines_absolute_"+str(deltaframes))==False:    
     os.mkdir(os.path.join(parent_dir, "Outlines_absolute_"+str(deltaframes)) )  


print ("making images...")
image_st_list=[]    #for stack later
image_st_c_list=[]
image_abs_list=[]
image_id_list=[]


for file in sorted(glob.glob(image_j2_dir))[deltaframes:]:
     print (file)
     index = (file[-7:-4])
     df4 = pd.read_excel(parent_dir+"/Outlines_matched_"+str(deltaframes)+"frames.xlsx" , sheet_name="Outline_t"+index)
 
     read = io.imread(file)
     image_dLength_absolute = np.zeros(read.shape, dtype=np.uint16)
     image_dLength_strain = np.zeros(read.shape, dtype=np.uint16) 
     image_dLength_strain_capped = np.zeros(read.shape, dtype=np.uint16) 
     image_Id1=np.zeros(read.shape, dtype=np.uint16)   
     image_Id2=np.zeros(read.shape, dtype=np.uint16) 

     for i in range(len(df4["centroids2"])): 
          junction = []

#          print ("Junction ", i, "Bitmap: ", df4["Lookup table 16Bit strain capped"][i])
                   
          coords1=tup(df4["coords1"][i])
          coords2=tup(df4["coords2"][i]) 
#          print (coords1)
          for j in coords2:
#               print ((j[0], j[1]), df4["Lookup table 16Bit absolute"][i])
               image_dLength_absolute[(j[0], j[1])] = df4["Lookup table 16Bit absolute"][i]
               image_dLength_strain[(j[0], j[1])] = df4["Lookup table 16Bit strain"][i]  
               image_dLength_strain_capped[(j[0], j[1])] = df4["Lookup table 16Bit strain capped"][i]
               image_Id2[(j[0], j[1])] = i
          for j in coords1:
              image_Id1[(j[0], j[1])] = i
     imwrite(parent_dir+"/Outlines_absolute_"+str(deltaframes)+"/Absolute_"+index+".tif", image_dLength_absolute)
     imwrite(parent_dir+"/Outlines_strain_"+str(deltaframes)+"/Strain_"+index+".tif", image_dLength_strain)
     imwrite(parent_dir+"/Outlines_strain_capped"+str(cap)+"_"+str(deltaframes)+"/Strain_capped"+str(cap)+"_"+index+".tif", image_dLength_strain_capped)
     image_abs_list.append(image_dLength_absolute)
     image_st_list.append(image_dLength_strain)
     image_st_c_list.append(image_dLength_strain_capped)    
     image_stack=np.stack([image_Id1, image_Id2], axis=0)
     with TiffWriter(parent_dir+"/Outlines_Id_"+str(deltaframes)+"/Id_"+index+".tif", imagej=True) as tif:
         tif.save(image_stack)
     image_id_list.append(image_stack)
print ("Folders with images of Junction ID, Absolute differences and Strain are saved")


#making stacks
print ("making stacks...")
image_st_stack = np.stack(image_st_list, axis=0)
with TiffWriter(parent_dir+"/Outlines_Strain_"+str(deltaframes)+"frames_stack.tif", imagej=True) as tif:
    tif.save(image_st_stack)
image_st_c_stack = np.stack(image_st_c_list, axis=0)
with TiffWriter(parent_dir+"/Outlines_Strain_capped"+str(cap)+"_"+str(deltaframes)+"frames_stack.tif", imagej=True) as tif:
    tif.save(image_st_c_stack)
image_abs_stack = np.stack(image_abs_list, axis=0)
with TiffWriter(parent_dir+"/Outlines_Absolute_"+str(deltaframes)+"frames_stack.tif", imagej=True) as tif:
    tif.save(image_abs_stack)
image_id_stack = np.stack(image_id_list, axis=0)
with TiffWriter(parent_dir+"/Outlines_Id_"+str(deltaframes)+"frames_stack.tif", imagej=True) as tif:
    tif.save(image_id_stack)    
    print ("Images created: Outlines_Strain_stack.tif, Outlines_Absolute_stack.tif, Outlines_Id_stack.tif")

